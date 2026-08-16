#!/usr/bin/env python3
"""Canonical Lacks lineage check (Codex, PR #17 final review).

The roadmap's claim — copy is AUTHORED in incoming/lacks_store_values.json
(and its sibling build inputs), REBUILT through incoming/build_lacks_workbook
.py, and EMITTED by tools/convert_store_data.py — was documentation, not a CI
guarantee: workbook validation checked the committed xlsx and strict golden
exercised the Bel fixture, but nothing proved the committed Lacks workbook and
bundle still DERIVE from the committed sources. A hand edit to the workbook or
a source edit without a rebuild would have shipped silently.

This check makes the chain executable, entirely in temp space (no committed
artifact is touched):

  1. Rebuild the workbook from the committed incoming JSON sources
     (build_lacks_workbook.py --out <tmp>).
  2. Compare it SEMANTICALLY with incoming/Lacks_Store_Data.xlsx — sheet
     names, order, and every cell value. Raw xlsx bytes are deliberately not
     compared: openpyxl restamps zip/docProps timestamps on every save, so
     byte equality is impossible and meaningless; cell equality is the truth.
  3. Validate the rebuilt workbook with warnings-as-errors (the same gate CI
     applies to the committed one).
  4. Convert the rebuilt workbook in temp space and canonically compare every
     applicable generated output with the committed artifact (parse-based,
     via tests/golden/canonical.py — data/mattresses.json is derived from the
     compared CSVs by build-data.ps1 and is covered by strict golden).
  5. NON-VACUITY: copy the sources to a sandbox, mutate one implication
     string, rebuild + convert there, and require BOTH observers (workbook
     compare, bundle compare) to turn red. A lineage check that cannot detect
     a source edit proves nothing.

A source change without a rebuild fails step 2/4; a workbook or bundle hand
edit fails the same steps from the other side.

Run: python tests/lineage_check.py
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(HERE, "golden"))
import canonical  # noqa: E402

BUILDER = os.path.join(REPO, "incoming", "build_lacks_workbook.py")
CONVERTER = os.path.join(REPO, "tools", "convert_store_data.py")
VALIDATOR = os.path.join(REPO, "tools", "validate_workbook.py")
COMMITTED_WB = os.path.join(REPO, "incoming", "Lacks_Store_Data.xlsx")

passed = failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
        print(f"  [ok] {name}")
    else:
        failed += 1
        print(f"  [FAIL] {name}")
        if detail:
            print(f"         {detail[:600]}")


def run(cmd, **kw):
    p = subprocess.run(cmd, capture_output=True, text=True, **kw)
    return p.returncode, p.stdout + p.stderr


def load_cells(path):
    """{sheet: [[cell,...],...]} with None normalized to '' and trailing empty
    cells/rows trimmed — the builder writes '' for absent optional cells, and
    openpyxl may round-trip those as None."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = []
        for row in wb[name].iter_rows(values_only=True):
            cells = ["" if v is None else v for v in row]
            while cells and cells[-1] == "":
                cells.pop()
            rows.append(cells)
        while rows and not rows[-1]:
            rows.pop()
        out[name] = rows
    wb.close()
    return out


def diff_workbooks(a_path, b_path):
    """Semantic differences between two workbooks; [] when equal."""
    a, b = load_cells(a_path), load_cells(b_path)
    diffs = []
    if list(a) != list(b):
        diffs.append(f"sheet names/order differ: {list(a)} != {list(b)}")
        return diffs
    for name in a:
        ra, rb = a[name], b[name]
        if len(ra) != len(rb):
            diffs.append(f"{name}: row count {len(ra)} != {len(rb)}")
        for i, (xa, xb) in enumerate(zip(ra, rb), start=1):
            if xa != xb:
                diffs.append(f"{name} row {i}: {str(xa)[:80]!r} != {str(xb)[:80]!r}")
            if len(diffs) >= 12:
                return diffs
    return diffs


def parse_allowed_hosts(path):
    import re
    m = re.search(r"__DF_ALLOWED_HOSTS\s*=\s*(\[.*?\])\s*;",
                  open(path, encoding="utf-8").read(), re.DOTALL)
    return json.loads(m.group(1)) if m else None


def compare_bundle(bundle_dir, label):
    """Canonically compare a converted bundle against the committed artifacts.
    Returns the list of per-file difference summaries ([] when identical)."""
    diffs = []
    for rel in ("data/store-config.json", "data/accessories.json",
                "data/quiz.json", "manifest.json"):
        got = os.path.join(bundle_dir, rel.replace("/", os.sep))
        want = os.path.join(REPO, rel.replace("/", os.sep))
        res = canonical.compare_json_files(want, got, label=f"{label}:{rel}")
        if not res:
            diffs.append(f"{rel}: {res.summary()[:200]}")
    for rel in ("data/mattresses.csv", "data/mattresses-es.csv"):
        got = os.path.join(bundle_dir, rel.replace("/", os.sep))
        want = os.path.join(REPO, rel.replace("/", os.sep))
        res = canonical.compare_csv_files(want, got, label=f"{label}:{rel}")
        if not res:
            diffs.append(f"{rel}: {res.summary()[:200]}")
    ah_got = parse_allowed_hosts(os.path.join(bundle_dir, "data", "allowed-hosts.js"))
    ah_want = parse_allowed_hosts(os.path.join(REPO, "data", "allowed-hosts.js"))
    if ah_got != ah_want:
        diffs.append(f"allowed-hosts.js: {ah_got} != {ah_want}")
    return diffs


def convert(workbook, out_dir):
    return run([sys.executable, CONVERTER, workbook,
                "--output-dir", out_dir, "--skip-build-json"])


def main():
    tmp = tempfile.mkdtemp(prefix="df-lineage-")
    try:
        print("-- 1+2: sources -> workbook, compared semantically with committed --")
        rebuilt = os.path.join(tmp, "rebuilt.xlsx")
        rc, out = run([sys.executable, BUILDER, "--out", rebuilt])
        check("builder rebuilds from the committed incoming sources", rc == 0, out)
        check("committed workbook untouched by the rebuild",
              os.path.exists(COMMITTED_WB))
        diffs = diff_workbooks(COMMITTED_WB, rebuilt)
        check("rebuilt workbook is cell-identical to incoming/Lacks_Store_Data.xlsx",
              not diffs, "; ".join(diffs))

        print("-- 3: the rebuilt workbook passes the same gate CI applies --")
        rc, out = run([sys.executable, VALIDATOR, rebuilt,
                       "--source-images", os.path.join(REPO, "incoming", "images"),
                       "--warnings-as-errors"])
        check("rebuilt workbook validates (warnings-as-errors)", rc == 0, out[-600:])

        print("-- 4: rebuilt workbook -> bundle, compared canonically with committed --")
        bundle = os.path.join(tmp, "bundle")
        rc, out = convert(rebuilt, bundle)
        check("converter emits a bundle from the rebuilt workbook", rc == 0, out[-600:])
        diffs = compare_bundle(bundle, "lineage")
        check("every generated artifact matches the committed one canonically",
              not diffs, "; ".join(diffs))

        print("-- 5: NON-VACUITY - a mutated source must turn both observers red --")
        sandbox = os.path.join(tmp, "sandbox")
        os.makedirs(sandbox)
        shutil.copytree(os.path.join(REPO, "incoming"), os.path.join(sandbox, "incoming"))
        shutil.copytree(os.path.join(REPO, "tools"), os.path.join(sandbox, "tools"))
        values_path = os.path.join(sandbox, "incoming", "lacks_store_values.json")
        values = json.load(open(values_path, encoding="utf-8"))
        impl = values["salesNotes.consultationImplications"]
        assert impl["health_conditions"]["snoring"]["en"], "probe target missing"
        impl["health_conditions"]["snoring"]["en"] = "LINEAGE-MUTATION-SENTINEL"
        with open(values_path, "w", encoding="utf-8") as f:
            json.dump(values, f, ensure_ascii=False, indent=2)
        mutated_wb = os.path.join(sandbox, "incoming", "Lacks_Store_Data.xlsx")
        rc, out = run([sys.executable,
                       os.path.join(sandbox, "incoming", "build_lacks_workbook.py"),
                       "--out", mutated_wb])
        check("sandboxed builder rebuilds from the mutated source", rc == 0, out[-400:])
        wb_diffs = diff_workbooks(COMMITTED_WB, mutated_wb)
        check("workbook observer turns RED on the mutated implication",
              any("LINEAGE-MUTATION-SENTINEL" in d or "SalesNotes" in d
                  for d in wb_diffs))
        mut_bundle = os.path.join(sandbox, "bundle")
        rc, out = convert(mutated_wb, mut_bundle)
        check("converter accepts the mutated (still-valid) workbook", rc == 0, out[-400:])
        bundle_diffs = compare_bundle(mut_bundle, "mutated")
        check("bundle observer turns RED on the mutated implication",
              any("store-config.json" in d for d in bundle_diffs))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print(f"\nLineage check: {passed} passed, {failed} failed")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
