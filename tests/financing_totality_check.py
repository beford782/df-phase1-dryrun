#!/usr/bin/env python3
"""Financing validation totality check.

validate_financing() is a BUILD GATE: the converter refuses to publish a bundle
when it reports errors. That only works if it always produces a verdict. A
traceback is not a refusal an operator can act on, and the build cannot tell it
apart from a validator defect — so for every value python's json.loads can
produce, anywhere in the financing subtree, this must hold:

  1. no exception escapes;
  2. a ValidationReport is returned;
  3. malformed input produces at least one error naming the offending field;
  4. no diagnostic is unbounded (a 5,000-digit id must not become the message);
  5. the caller's config object is not mutated.

json.loads can produce: object, array, string, integer (unbounded), float
(including NaN / Infinity / -Infinity, which it accepts by default), boolean
and null. It cannot produce tuples, sets, bytes or subclasses, so those are out
of scope by construction.

Run: python tests/financing_totality_check.py     (exit 0 = all pass)
"""
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import traceback

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO, "tools"))
import validation  # noqa: E402

passed = failed = 0

# Longest tolerable single error string. The static prose in the taxonomy and
# ungated-copy errors is already a few hundred characters; anything beyond this
# means an author value was interpolated without bounds.
MAX_ERROR_CHARS = 900
# A long digit run is the signature of a huge integer reaching the message.
LONG_DIGIT_RUN = re.compile(r"\d{200,}")

HUGE = 10 ** 5000
NAN, INF, NINF = float("nan"), float("inf"), float("-inf")

# Every shape json.loads can hand us. Order matters only for readability.
JSON_VALUES = [
    ("null", None), ("true", True), ("false", False),
    ("zero", 0), ("negative", -1), ("float", 1.5), ("huge int", HUGE),
    ("NaN", NAN), ("Infinity", INF), ("-Infinity", NINF),
    ("empty string", ""), ("string", "bad"),
    ("empty array", []), ("array", ["x"]), ("nested array", [[["deep"]]]),
    ("empty object", {}), ("object", {"a": 1}),
    ("nested object", {"a": {"b": {"c": 1}}}),
]


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
    else:
        failed += 1
        print(f"  [FAIL] {name}" + (f" — {detail}" if detail else ""))
    return cond


def load(rel):
    with io.open(os.path.join(REPO, rel), encoding="utf-8") as f:
        return json.load(f)


HOSTS = load("tools/source_hosts.json")["financingSourceHosts"]
BASE = load("data/store-config.json")


def _fingerprint(o):
    """Reduce a document to something json.dumps can always encode. A huge
    integer is described rather than rendered — json.dumps hits the very same
    int->str digit limit this suite exists to police, so the harness would
    otherwise crash on the inputs it is meant to check."""
    if isinstance(o, dict):
        return {k: _fingerprint(v) for k, v in o.items()}
    if isinstance(o, list):
        return [_fingerprint(v) for v in o]
    if isinstance(o, int) and not isinstance(o, bool) and o.bit_length() > 128:
        return f"<int:{o.bit_length()} bits>"
    if isinstance(o, float) and (o != o or o in (INF, NINF)):
        return f"<float:{o!r}>"
    return o


def stable(obj):
    """Order-independent, NaN-tolerant fingerprint for mutation detection."""
    return json.dumps(_fingerprint(obj), sort_keys=True, allow_nan=True,
                      default=repr)


def probe(label, build, expect_field=None):
    """Run validate_financing and assert the five totality obligations.

    `build` may be a config or a callable producing one. Any exception raised
    while PREPARING a case is reported as a failure rather than aborting the
    run, so one bad fixture cannot hide the rest of the matrix."""
    try:
        config = build() if callable(build) else build
        before = stable(config)
    except BaseException as exc:                                  # noqa: BLE001
        check(f"{label}: harness prepared the case", False,
              f"{type(exc).__name__}: {str(exc)[:80]}")
        return None
    try:
        rep = validation.validate_financing(config, allowed_source_hosts=HOSTS)
    except BaseException as exc:                                  # noqa: BLE001
        frames = [f for f in traceback.extract_tb(exc.__traceback__)
                  if "validation.py" in (f.filename or "")]
        site = f" at validation.py:{frames[-1].lineno}" if frames else ""
        check(f"{label}: returns a report (no exception)", False,
              f"{type(exc).__name__}: {str(exc)[:70]}{site}")
        return None
    ok = check(f"{label}: returns a ValidationReport",
               isinstance(rep, validation.ValidationReport))
    try:
        longest = max((len(e) for e in rep.errors + rep.warnings), default=0)
    except BaseException as exc:                                  # noqa: BLE001
        check(f"{label}: diagnostics are readable strings", False,
              f"{type(exc).__name__}: {exc}")
        return rep
    ok &= check(f"{label}: diagnostics bounded", longest <= MAX_ERROR_CHARS,
                f"longest={longest}")
    ok &= check(f"{label}: no huge value embedded in a diagnostic",
                not any(LONG_DIGIT_RUN.search(e) for e in rep.errors + rep.warnings))
    ok &= check(f"{label}: input not mutated", stable(config) == before)
    if expect_field is not None:
        ok &= check(f"{label}: names the offending field {expect_field!r}",
                    any(expect_field in e for e in rep.errors),
                    "; ".join(e[:70] for e in rep.errors[:2]) or "(no errors)")
    return rep


def cfg_with(path, value):
    """Deep copy of the shipped config with one path replaced. `path` is a
    sequence of dict keys / list indices; () replaces the whole document."""
    doc = json.loads(json.dumps(BASE))
    if not path:
        return value
    node = doc
    for key in path[:-1]:
        node = node[key]
    node[path[-1]] = value
    return doc


def main():
    print("Top-level and block shapes:")
    for lbl, val in JSON_VALUES:
        probe(f"config = {lbl}", cfg_with((), val),
              None if isinstance(val, dict) else "config must be an object")
    for lbl, val in JSON_VALUES:
        expect = None if val is None or isinstance(val, dict) else "financing"
        probe(f"financing = {lbl}", cfg_with(("financing",), val), expect)
    for lbl, val in JSON_VALUES:
        expect = None if val is None or isinstance(val, dict) else "discount"
        probe(f"discount = {lbl}", cfg_with(("discount",), val), expect)

    print("financing.copy and its leaves:")
    for lbl, val in JSON_VALUES:
        expect = None if val is None or isinstance(val, dict) else "financing.copy"
        probe(f"copy = {lbl}", cfg_with(("financing", "copy"), val), expect)
    for lbl, val in JSON_VALUES:
        # A plain string is legal single-language copy; a bilingual object is
        # legal; everything else must be named.
        expect = None if isinstance(val, str) else "financing.copy.body"
        probe(f"copy.body = {lbl}",
              cfg_with(("financing", "copy", "body"), val), expect)
    for lbl, val in JSON_VALUES:
        expect = None if isinstance(val, str) and val.strip() else "financing.copy.body"
        probe(f"copy.body.en = {lbl}",
              cfg_with(("financing", "copy", "body"), {"en": val, "es": "ok"}),
              expect)
        probe(f"copy.body.es = {lbl}",
              cfg_with(("financing", "copy", "body"), {"en": "ok", "es": val}),
              expect)

    print("financing.plans (the collection):")
    for lbl, val in JSON_VALUES:
        expect = None if val is None or isinstance(val, list) else "financing.plans"
        probe(f"plans = {lbl}", cfg_with(("financing", "plans"), val), expect)
    print("financing.plans entries:")
    for lbl, val in JSON_VALUES:
        expect = None if isinstance(val, dict) else "financing.plans[0]"
        probe(f"plans[0] = {lbl}", cfg_with(("financing", "plans"), [val]), expect)
        probe(f"plans[+] = {lbl} (appended to valid plans)",
              cfg_with(("financing", "plans"),
                       json.loads(json.dumps(BASE))["financing"]["plans"] + [val]),
              None if isinstance(val, dict) else "financing.plans[6]")

    print("plan fields:")
    PLAN_FIELDS = [
        ("id", lambda v: not (isinstance(v, str) and v.strip())),
        ("kind", lambda v: True),
        ("provider", lambda v: not (isinstance(v, str) and v.strip() == v and v.strip())),
        ("presentationScenario", lambda v: True),
        ("headline", lambda v: True),
        ("detail", lambda v: True),
        ("disclosure", lambda v: True),
        ("sourceUrl", lambda v: True),
        ("verifiedAt", lambda v: True),
        ("apr", lambda v: True),
        ("termMonths", lambda v: True),
        ("minimumPurchase", lambda v: True),
        ("publishedPaymentFactor", lambda v: True),
    ]
    for field, _bad in PLAN_FIELDS:
        for lbl, val in JSON_VALUES:
            probe(f"plans[0].{field} = {lbl}",
                  cfg_with(("financing", "plans", 0, field), val))
    print("bilingual leaves of plan copy:")
    for field in ("headline", "detail", "disclosure"):
        for lbl, val in JSON_VALUES:
            probe(f"plans[0].{field}.en = {lbl}",
                  cfg_with(("financing", "plans", 0, field), {"en": val, "es": "ok"}))

    print("financing scalars and collections:")
    for field in ("enabled", "experience", "offerVersion", "verifiedAt", "maxAgeDays",
                  "sourceUrl", "applicationUrl", "mexicoInfoUrl", "mexicoApplicationUrl",
                  "savingsPassPolicy", "exactPromotionsEnabled", "esReviewStatus",
                  "allowedSourceHosts"):
        for lbl, val in JSON_VALUES:
            probe(f"financing.{field} = {lbl}",
                  cfg_with(("financing", field), val))
    for lbl, val in JSON_VALUES:
        probe(f"allowedSourceHosts[0] = {lbl}",
              cfg_with(("financing", "allowedSourceHosts"), [val]),
              None if isinstance(val, str) else "allowedSourceHosts")
        probe(f"mexicoApplicationUrl.url = {lbl}",
              cfg_with(("financing", "mexicoApplicationUrl"),
                       {"url": val, "verified": False}))

    print("Named regressions - the shapes that used to raise:")
    NAMED = [
        ("plans = ['bad']", ("financing", "plans"), ["bad"], "financing.plans[0]"),
        ("plans = 'bad'", ("financing", "plans"), "bad", "financing.plans"),
        ("copy = 'bad'", ("financing", "copy"), "bad", "financing.copy"),
        ("plan.kind = []", ("financing", "plans", 0, "kind"), [], "kind"),
        ("discount = 'bad'", ("discount",), "bad", "discount"),
        ("plan id = huge int", ("financing", "plans", 0, "id"), HUGE, "id"),
        ("headline.en = huge int", ("financing", "plans", 0, "headline"),
         {"en": HUGE, "es": "x"}, "headline"),
        ("minimumPurchase = true", ("financing", "plans", 0, "minimumPurchase"),
         True, "minimumPurchase"),
        ("minimumPurchase = NaN", ("financing", "plans", 0, "minimumPurchase"),
         NAN, "minimumPurchase"),
        ("minimumPurchase = Infinity", ("financing", "plans", 0, "minimumPurchase"),
         INF, "minimumPurchase"),
        ("minimumPurchase = -Infinity", ("financing", "plans", 0, "minimumPurchase"),
         NINF, "minimumPurchase"),
        ("maxAgeDays = true", ("financing", "maxAgeDays"), True, "maxAgeDays"),
        ("publishedPaymentFactor = true",
         ("financing", "plans", 0, "publishedPaymentFactor"), True,
         "publishedPaymentFactor"),
        ("savingsPassPolicy = []", ("financing", "savingsPassPolicy"), [],
         "savingsPassPolicy"),
        ("verifiedAt = huge int", ("financing", "verifiedAt"), HUGE, "verifiedAt"),
    ]
    for lbl, path, val, expect in NAMED:
        probe(lbl, cfg_with(path, val), expect)

    print("Valid-input verdicts are preserved:")
    shipped = validation.validate_financing(json.loads(json.dumps(BASE)),
                                            allowed_source_hosts=HOSTS)
    check("the shipped config still validates clean", shipped.ok,
          "; ".join(shipped.errors[:2]))
    check("absent financing is still a no-op",
          validation.validate_financing({}, allowed_source_hosts=HOSTS).ok)
    check("null financing is still a no-op",
          validation.validate_financing({"financing": None},
                                        allowed_source_hosts=HOSTS).ok)
    for good, term in ((0, 1), (100, 120), (9.99, 72), (0.0, 48)):
        c = cfg_with(("financing", "plans", 0, "apr"), good)
        c["financing"]["plans"][0]["termMonths"] = term
        import financing_headline as fh
        c["financing"]["plans"][0]["headline"] = fh.promotional_headline(good, term)
        rep = validation.validate_financing(c, allowed_source_hosts=HOSTS)
        check(f"in-range apr={good!r}/term={term} still validates clean", rep.ok,
              "; ".join(rep.errors[:1]))
    for mp in (0, 1, 500, 4200, 0.5, 99999):
        rep = validation.validate_financing(
            cfg_with(("financing", "plans", 0, "minimumPurchase"), mp),
            allowed_source_hosts=HOSTS)
        check(f"ordinary minimumPurchase {mp!r} still accepted",
              not any("minimumPurchase" in e for e in rep.errors))

    print("Taxonomy helpers are total:")
    for lbl, val in JSON_VALUES:
        for helper in (validation._plan_group, validation._plan_scenario,
                       validation._ungated_plan_fields,
                       validation._is_gated_offer_plan):
            try:
                helper(val)
                helper({"kind": val})
                helper({"presentationScenario": val})
                check(f"{helper.__name__} total for {lbl}", True)
            except BaseException as exc:                          # noqa: BLE001
                check(f"{helper.__name__} total for {lbl}", False,
                      f"{type(exc).__name__}: {exc}")
    check("valid taxonomy unchanged: promotional",
          validation._plan_group({"kind": "open-end-promotional-credit"}) == "promotional")
    check("valid taxonomy unchanged: installment",
          validation._plan_group({"kind": "closed-end-installment"}) == "installment")
    check("valid taxonomy unchanged: evergreen",
          validation._plan_group({"kind": "lease-to-own"}) == "evergreen")
    check("valid taxonomy unchanged: scenario",
          validation._plan_group({"kind": "closed-end-installment",
                                  "presentationScenario": "mexico-delivery"}) == "scenario")

    print("The allowlist ARGUMENT is JSON too (tools/source_hosts.json):")
    # load_source_hosts() reads an editable JSON file and hands the result
    # straight in, so a malformed allowlist must fail closed with a verdict,
    # not raise. It also must not silently widen: entries that are not host
    # strings are dropped AND reported.
    for lbl, val in JSON_VALUES + [("list with non-string entries", [5, None, []]),
                                   ("mixed list", ["lacks.com", 7])]:
        cfg = json.loads(json.dumps(BASE))
        try:
            rep = validation.validate_financing(cfg, allowed_source_hosts=val)
            check(f"allowed_source_hosts = {lbl}: returns a report",
                  isinstance(rep, validation.ValidationReport))
            check(f"allowed_source_hosts = {lbl}: diagnostics bounded",
                  all(len(e) <= MAX_ERROR_CHARS for e in rep.errors))
        except BaseException as exc:                              # noqa: BLE001
            check(f"allowed_source_hosts = {lbl}: returns a report", False,
                  f"{type(exc).__name__}: {str(exc)[:70]}")
    check("a malformed allowlist fails CLOSED (no URL passes)",
          any("must be a safe https" in e or "allowlisted host" in e
              for e in validation.validate_financing(
                  json.loads(json.dumps(BASE)), allowed_source_hosts=5).errors))
    check("a malformed allowlist is reported, not silently tolerated",
          any("allowed_source_hosts" in e for e in validation.validate_financing(
              json.loads(json.dumps(BASE)), allowed_source_hosts=5).errors))
    check("non-string allowlist entries are reported",
          any("non-string entr" in e for e in validation.validate_financing(
              json.loads(json.dumps(BASE)),
              allowed_source_hosts=["lacks.com", 7]).errors))
    check("_is_allowed_source is total over allowlist shapes",
          all(validation._is_allowed_source("https://www.lacks.com/financing", v)
              is not None for _l, v in JSON_VALUES))
    check("the real allowlist still admits the shipped source URL",
          validation._is_allowed_source("https://www.lacks.com/financing", HOSTS))

    print("Long strings do not produce unbounded diagnostics:")
    long_text = "x" * 200000
    for path in (("financing", "applicationUrl"), ("financing", "sourceUrl"),
                 ("financing", "verifiedAt"), ("financing", "savingsPassPolicy"),
                 ("financing", "plans", 0, "id")):
        probe(f"{'.'.join(str(p) for p in path)} = 200k-char string",
              cfg_with(path, long_text))
    probe("plans[0].headline.en = 200k-char string",
          cfg_with(("financing", "plans", 0, "headline"),
                   {"en": long_text, "es": long_text}))

    print("Converter fixtures (the real build gate, not a helper call):")
    _converter_fixtures()

    print(f"\nFinancing totality check: {passed} passed, {failed} failed")
    return 1 if failed else 0


def _write_hostile_workbook(dst, mutate):
    """Copy the committed workbook and rewrite its Promotions envelope."""
    import openpyxl
    shutil.copy2(os.path.join(REPO, "incoming", "Lacks_Store_Data.xlsx"), dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Promotions"]
    payload = "".join((c.value or "") for (c,) in ws.iter_rows(min_row=2, max_col=1))
    env = json.loads(payload)
    mutate(env["financing"])
    new = json.dumps(env, ensure_ascii=False, separators=(",", ":"))
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    for i in range(0, len(new), 30000):
        ws.append([new[i:i + 30000]])
    wb.save(dst)


def _write_hostile_envelope(dst, financing_value):
    """Like _write_hostile_workbook, but replaces the whole financing value —
    including falsy shapes the envelope reader used to drop."""
    import openpyxl
    shutil.copy2(os.path.join(REPO, "incoming", "Lacks_Store_Data.xlsx"), dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Promotions"]
    payload = "".join((c.value or "") for (c,) in ws.iter_rows(min_row=2, max_col=1))
    env = json.loads(payload)
    env["financing"] = financing_value
    new = json.dumps(env, ensure_ascii=False, separators=(",", ":"))
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    for i in range(0, len(new), 30000):
        ws.append([new[i:i + 30000]])
    wb.save(dst)


def _converter_fixtures():
    """A malformed workbook must make the converter refuse CLEANLY: nonzero
    exit, a named error, and no traceback — in --validate-only mode and in the
    normal publishing mode, which must refuse before writing anything."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        check("converter fixtures (openpyxl available)", False, "openpyxl missing")
        return
    cases = [
        ("plans = 'bad'", lambda f: f.__setitem__("plans", "bad"), "financing.plans"),
        ("plans[0] = 'bad'", lambda f: f.__setitem__("plans", ["bad"]),
         "financing.plans[0]"),
        ("copy = 'bad'", lambda f: f.__setitem__("copy", "bad"), "financing.copy"),
        ("plan.kind = []", lambda f: f["plans"][0].__setitem__("kind", []), "kind"),
        ("minimumPurchase = Infinity",
         lambda f: f["plans"][0].__setitem__("minimumPurchase", INF),
         "minimumPurchase"),
    ]
    # A present-but-FALSY financing envelope used to be dropped by the
    # converter's `if financing:` test, so validate_financing never saw it and
    # the build published a bundle with Payment Choice silently missing.
    falsy = [("financing = []", []), ("financing = ''", ""),
             ("financing = 0", 0), ("financing = false", False)]
    tmp = tempfile.mkdtemp(prefix="df_totality_")
    try:
        for lbl, mutate, expect in cases:
            wb_path = os.path.join(tmp, "hostile.xlsx")
            _write_hostile_workbook(wb_path, mutate)
            for mode, extra in (("--validate-only", ["--validate-only"]),
                                ("publishing", ["--output-dir",
                                                os.path.join(tmp, "out")])):
                proc = subprocess.run(
                    [sys.executable, os.path.join(REPO, "tools", "convert_store_data.py"),
                     wb_path] + extra,
                    capture_output=True, text=True, cwd=REPO)
                out = (proc.stdout or "") + (proc.stderr or "")
                check(f"converter {mode}: {lbl} exits nonzero", proc.returncode != 0,
                      f"rc={proc.returncode}")
                check(f"converter {mode}: {lbl} prints no traceback",
                      "Traceback" not in out, out.strip()[-120:])
                check(f"converter {mode}: {lbl} names the field",
                      expect in out, out.strip()[-120:])
            # Publishing mode must refuse BEFORE writing the bundle.
            check(f"converter publishing: {lbl} wrote no store-config",
                  not os.path.exists(os.path.join(tmp, "out", "data",
                                                  "store-config.json")))
        for lbl, val in falsy:
            wb_path = os.path.join(tmp, "falsy.xlsx")
            _write_hostile_envelope(wb_path, val)
            proc = subprocess.run(
                [sys.executable, os.path.join(REPO, "tools", "convert_store_data.py"),
                 wb_path, "--validate-only"],
                capture_output=True, text=True, cwd=REPO)
            out = (proc.stdout or "") + (proc.stderr or "")
            check(f"converter refuses a falsy envelope: {lbl}",
                  proc.returncode != 0 and "financing" in out,
                  f"rc={proc.returncode} {out.strip()[-100:]}")
            check(f"converter prints no traceback: {lbl}", "Traceback" not in out)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
